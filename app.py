from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file, abort
from database import Database
from functools import wraps
import os, io, re, json, hashlib, secrets, time, threading, webbrowser
from datetime import datetime, date, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import sys

# ------------------------------------------------------------
#  إعداد المفتاح السرّي بشكل ثابت (لا يتغير مع كل تشغيل)
# ------------------------------------------------------------
def _get_secret_key():
    # 1) من متغير البيئة
    env = os.environ.get('PAYROLL_SECRET_KEY')
    if env:
        return env
    # تحديد المجلد الأساسي: بجوار الـ exe في وضع PyInstaller، أو بجانب app.py
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    keyfile = os.path.join(base_dir, '.secret_key')
    # 2) من ملف سري - أول شركة
    if os.path.exists(keyfile):
        with open(keyfile, 'r', encoding='utf-8') as f:
            v = f.read().strip()
            if v:
                return v
    # 3) إنشاء وتخزين
    v = secrets.token_hex(32)
    try:
        with open(keyfile, 'w', encoding='utf-8') as f:
            f.write(v)
    except Exception:
        pass
    return v

app = Flask(__name__)
app.secret_key = _get_secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    # فُعّل HTTPS (SESSION_COOKIE_SECURE=True) عبر متغير البيئة PAYROLL_HTTPS=1 عند النشر الخارجي
    SESSION_COOKIE_SECURE=(os.environ.get('PAYROLL_HTTPS') == '1'),
    MAX_CONTENT_LENGTH=16 * 1024 * 1024,
)

db = Database()

# ملاحظة: مجلد القوالب والاستاتيك ضمن التطبيق (مجلد التحميل المؤقت في وضع الـ exe)
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
    DATA_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DATA_DIR = os.path.dirname(os.path.abspath(__file__))

# ==================== HELPERS / SECURITY ====================

# تخزين مقيد للجلسة لمنع كسر التصعيد
@app.before_request
def before_request():
    db._uid = session.get('user_id')
    db._uname = session.get('username','')
    db._ip = request.remote_addr or ''
    # إلزام المدير بتغيير كلمة المرور الافتراضية قبل استخدام باقي النظام
    if 'user_id' in session and request.endpoint not in ('static','logout','change_password','password_page'):
        u = db.get_user_by_id(session['user_id'])
        if u and db.admin_must_change_password(u):
            return redirect(url_for('password_page'))
    # نسخة احتياطية تلقائية شهرية (مرة كل شهر)
    if session.get('role') == 'admin':
        db.run_auto_backup_if_due()

@app.after_request
def security_headers(resp):
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    resp.headers['X-Frame-Options'] = 'SAMEORIGIN'
    resp.headers['X-XSS-Protection'] = '1; mode=block'
    resp.headers['Referrer-Policy'] = 'same-origin'
    # منع التخزين المؤقت للكشوفات ودخول السجلات الحساسة
    if request.endpoint in ('payslip','reports','dept_report','annual_report','login',
                            'comparison','signatures_log','audit','employees','contracts'):
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
    return resp

def login_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if 'user_id' not in session:
            flash('يجب تسجيل الدخول أولاً', 'error')
            return redirect(url_for('login'))
        return f(*a, **kw)
    return wrapper

# تسلسل الصلاحيات: admin > supervisor > user > viewer
_ROLE_LEVEL = {'admin': 4, 'supervisor': 3, 'user': 2, 'viewer': 1}

def role_required(min_role):
    """حماية مسار بناءً على الحد الأدنى للصلاحية."""
    def decorator(f):
        @wraps(f)
        def wrapper(*a, **kw):
            if 'user_id' not in session:
                flash('يجب تسجيل الدخول أولاً', 'error')
                return redirect(url_for('login'))
            user = db.get_user_by_id(session['user_id'])
            user_level = _ROLE_LEVEL.get(user.get('role',''), 0) if user else 0
            if user_level < _ROLE_LEVEL.get(min_role, 99):
                flash('ليس لديك صلاحية كافية للوصول', 'error')
                return redirect(url_for('dashboard'))
            return f(*a, **kw)
        return wrapper
    return decorator

def admin_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = db.get_user_by_id(session['user_id'])
        if not user or user['role'] != 'admin':
            flash('ليس لديك صلاحية للوصول', 'error')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return wrapper

# ==================== CSRF protection ====================
def csrf_token():
    if 'csrf' not in session:
        session['csrf'] = secrets.token_hex(32)
    return session['csrf']

def csrf_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if request.method == 'POST':
            token = request.form.get('_csrf') or (request.get_json(silent=True) or {}).get('_csrf') or ''
            if not token or token != session.get('csrf'):
                abort(400, 'انتهت صلاحية الجلسة (CSRF)')
        return f(*a, **kw)
    return wrapper

@app.context_processor
def inject_globals():
    user = None
    if 'user_id' in session:
        user = db.get_user_by_id(session['user_id'])
    settings = db.get_settings()
    # تنبيهات العقود المنتهية قريباً (للمسؤول)
    expiring = db.get_expiring_contracts(30) if session.get('role') == 'admin' else []
    return dict(current_user=user, sys_settings=settings, today=date.today().isoformat(),
                current_month=datetime.now().strftime('%Y-%m'),
                csrf_token=csrf_token, expiring_contracts=expiring)

def month_range(month):
    y, m = int(month[:4]), int(month[5:7])
    return calendar.monthrange(y, m)[1]

def _t(template, **kw):
    kw['csrf_token'] = csrf_token()
    return render_template(template, **kw)

# ==================== AUTH ====================
# محاولات تسجيل دخول فاشلة بسيطة لمنع التخمين
_login_attempts = {}

@app.route('/login', methods=['GET','POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        ip = request.remote_addr or ''
        # حدّ من المحاولات
        now = datetime.now()
        for k in list(_login_attempts.keys()):
            if now - _login_attempts[k] > timedelta(minutes=15):
                del _login_attempts[k]
        key = f"{ip}|{username}"
        if _login_attempts.get(key, now - timedelta(minutes=15)) > now - timedelta(minutes=1):
            flash('محاولات كثيرة، انتظر قليلاً ثم حاول مرة أخرى', 'error')
            return render_template('login.html', csrf_token=csrf_token())
        user = db.authenticate(username, password)
        if user and user['active']:
            session.clear()
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            session['csrf'] = secrets.token_hex(32)
            db._uid = user['id']; db._uname = user['username']; db._ip = ip
            db.log_audit('login', 'user', user['id'], f'تسجيل دخول ناجح')
            # إلزام المدير الأولي بتغيير كلمة المرور الافتراضية
            if db.admin_must_change_password(user):
                flash('يجب تغيير كلمة المرور الافتراضية لأسباب أمنية', 'warning')
                return redirect(url_for('password_page'))
            return redirect(url_for('dashboard'))
        _login_attempts[key] = now
        db.log_audit('login_failed', 'user', None, f'محاولة دخول فاشلة: {username} من {ip}')
        flash('بيانات الدخول غير صحيحة', 'error')
    return render_template('login.html', csrf_token=csrf_token())

@app.route('/logout')
def logout():
    if 'user_id' in session:
        db._uid = session.get('user_id'); db._uname = session.get('username','')
        db.log_audit('logout', 'user', session.get('user_id'), 'تسجيل خروج')
    session.clear()
    return redirect(url_for('login'))

@app.route('/password', methods=['GET'])
@login_required
def password_page():
    mc = db.admin_must_change_password(db.get_user_by_id(session['user_id']))
    return render_template('change_password.html', must_change=mc)

@app.route('/change-password', methods=['POST'])
@login_required
@csrf_required
def change_password():
    user = db.get_user_by_id(session['user_id'])
    old = request.form.get('old_password') or ''
    new = request.form.get('new_password') or ''
    confirm = request.form.get('confirm_password') or ''
    must_change = db.admin_must_change_password(user)
    if not db.verify_password(user['password'], old):
        flash('كلمة المرور الحالية غير صحيحة', 'error')
    elif len(new) < 8:
        flash('كلمة المرور يجب أن تكون 8 أحرف على الأقل', 'error')
    elif new != confirm:
        flash('كلمتا المرور غير متطابقتين', 'error')
    else:
        db.change_password(user['id'], new)
        db.mark_password_changed()
        db.log_audit('change_password', 'user', user['id'], 'تغيير كلمة المرور')
        flash('تم تغيير كلمة المرور بنجاح', 'success')
        if must_change:
            return redirect(url_for('dashboard'))
    if must_change:
        return redirect(url_for('password_page'))
    return redirect(url_for('settings'))

# ==================== DASHBOARD ====================
@app.route('/')
@login_required
def dashboard():
    stats = db.get_dashboard_stats()
    month = datetime.now().strftime('%Y-%m')
    payroll = db.get_saved_payroll(month)
    dist = db.get_pay_type_distribution()
    budget = db.get_monthly_budget(month)
    return render_template('dashboard.html', stats=stats, payroll=payroll,
                           dist=dist, budget=budget, month=month)

# ==================== EMPLOYEES ====================
@app.route('/employees')
@login_required
def employees():
    dept = request.args.get('department')
    status = request.args.get('status','active')
    search = request.args.get('search','')
    sort = request.args.get('sort','name')
    emps = db.get_employees(dept, status, search, sort)
    depts = db.get_departments(True)
    return render_template('employees.html', emps=emps, depts=depts, dept=dept, status=status,
                           sort=sort, search=search)

@app.route('/employees/add', methods=['POST'])
@role_required('supervisor')
@csrf_required
def add_employee():
    d = request.form.to_dict()
    d['emp_code'] = (request.form.get('emp_code') or '').strip() or db.next_emp_code()
    if not (request.form.get('name') or '').strip():
        flash('اسم الموظف مطلوب', 'error')
        return redirect(url_for('employees'))
    for f in ['base_salary','daily_rate','hourly_rate','commission_rate','housing_allowance','transport_allowance',
              'food_allowance','other_allowances','danger_allowance','phone_allowance',
              'social_insurance_employee_ratio','social_insurance_employer_ratio','tax_exempt_amount']:
        d[f] = _num(request.form.get(f))
    d['social_insurance_enabled'] = 1 if request.form.get('social_insurance_enabled') else 0
    d['income_tax_enabled'] = 1 if request.form.get('income_tax_enabled') else 0
    new_id = db.add_employee(d)
    db.log_audit('add', 'employee', new_id, f'إضافة موظف: {d["name"]}')
    flash('تم إضافة الموظف بنجاح', 'success')
    return redirect(url_for('employees'))

@app.route('/employees/<int:emp_id>')
@login_required
def employee_detail(emp_id):
    emp = db.get_employee(emp_id)
    if not emp:
        abort(404)
    history = db.get_salary_history(emp_id)
    depts = db.get_departments(True)
    contracts = db.get_contracts(emp_id)
    return render_template('employee_detail.html', emp=emp, depts=depts, history=history, contracts=contracts)

@app.route('/employees/<int:emp_id>/edit', methods=['POST'])
@role_required('supervisor')
@csrf_required
def edit_employee(emp_id):
    old = db.get_employee(emp_id)
    if not old:
        abort(404)
    d = request.form.to_dict()
    for f in ['base_salary','daily_rate','hourly_rate','commission_rate','housing_allowance','transport_allowance',
              'food_allowance','other_allowances','danger_allowance','phone_allowance',
              'social_insurance_employee_ratio','social_insurance_employer_ratio','tax_exempt_amount']:
        d[f] = _num(request.form.get(f))
    d['social_insurance_enabled'] = 1 if request.form.get('social_insurance_enabled') else 0
    d['income_tax_enabled'] = 1 if request.form.get('income_tax_enabled') else 0
    db.update_employee(emp_id, d)
    if old and _num(old['base_salary']) != d['base_salary']:
        db.record_salary_change(emp_id, old['base_salary'], d['base_salary'],
                                request.form.get('salary_change_reason','تعديل الراتب'), session.get('username',''))
    db.log_audit('edit', 'employee', emp_id, f'تعديل بيانات موظف: {old["name"]}')
    flash('تم تحديث بيانات الموظف', 'success')
    return redirect(url_for('employee_detail', emp_id=emp_id))

@app.route('/employees/<int:emp_id>/delete', methods=['POST'])
@admin_required
@csrf_required
def delete_employee(emp_id):
    emp = db.get_employee(emp_id)
    db.delete_employee(emp_id)
    db.log_audit('delete', 'employee', emp_id, f'حذف موظف: {emp["name"] if emp else ""}')
    flash('تم حذف الموظف', 'success')
    return redirect(url_for('employees'))

@app.route('/employees/export')
@login_required
def export_employees():
    emps = db.get_employees(status='all')
    wb = Workbook()
    ws = wb.active
    ws.title = "الموظفين"
    headers = ['الكود','الاسم','الرقم القومي','القسم','الوظيفة','نوع الأجر','الراتب الأساسي','الراتب اليومي','سعر الساعة',
               'السكن','النقل','الطعام','أخرى','خطورة','موبايل','التأمينات','الضرائب','الحالة']
    ws.append(headers)
    for e in emps:
        ws.append([e['emp_code'],e['name'],e['national_id'],e.get('dept_name',''),e['position'],e['pay_type'],
                   e['base_salary'],e['daily_rate'],e['hourly_rate'],e['housing_allowance'],e['transport_allowance'],
                   e['food_allowance'],e['other_allowances'],e['danger_allowance'],e['phone_allowance'],
                   'نعم' if e['social_insurance_enabled'] else 'لا','نعم' if e['income_tax_enabled'] else 'لا',e['status']])
    # تنسيق فخم بأعمدة أرقام
    money = [7,8,9,10,11,12,13,14,15]  # الفهرس (1-base) للأعمدة المالية
    center = [1,4,6,16,17,18]
    _style_sheet(ws, 1, len(headers), money_cols=money, center_cols=center)
    out = _export_xlsx(wb)
    return send_file(out, as_attachment=True, download_name='الموظفين.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ==================== DEPARTMENTS ====================
@app.route('/departments')
@login_required
def departments():
    depts = db.get_departments()
    return render_template('departments.html', depts=depts)

@app.route('/departments/add', methods=['POST'])
@login_required
@csrf_required
def add_department():
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('اسم القسم مطلوب', 'error')
        return redirect(url_for('departments'))
    try:
        db.add_department(name, request.form.get('manager',''), request.form.get('cost_center',''))
        db.log_audit('add', 'department', None, f'إضافة قسم: {name}')
        flash('تم إضافة القسم', 'success')
    except Exception:
        flash('القسم موجود بالفعل', 'error')
    return redirect(url_for('departments'))

@app.route('/departments/<int:dep_id>/delete', methods=['POST'])
@admin_required
@csrf_required
def delete_department(dep_id):
    if not db.delete_department(dep_id):
        flash('لا يمكن حذف قسم به موظفين', 'error')
    else:
        db.log_audit('delete', 'department', dep_id, 'حذف قسم')
        flash('تم حذف القسم', 'success')
    return redirect(url_for('departments'))

# ==================== ATTENDANCE ====================
@app.route('/attendance')
@login_required
def attendance():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    emps = db.get_employees(status='active')
    records = db.get_attendance(month)
    return render_template('attendance.html', emps=emps, records=records, month=month)

@app.route('/attendance/save', methods=['POST'])
@role_required('user')
@csrf_required
def save_attendance():
    month = request.json.get('month')
    db.bulk_save_attendance(month, request.json.get('records', []))
    db.log_audit('attendance', 'attendance', None, f'تحديث حضور شهر {month}')
    return jsonify({'ok': True})

# ==================== LEAVES ====================
@app.route('/leaves')
@login_required
def leaves():
    emps = db.get_employees(status='active')
    records = db.get_leaves()
    return render_template('leaves.html', emps=emps, records=records)

@app.route('/leaves/add', methods=['POST'])
@role_required('user')
@csrf_required
def add_leave():
    d = request.form.to_dict()
    d['days'] = _num(request.form.get('days'))
    db.add_leave(d)
    db.log_audit('add', 'leave', None, 'إضافة طلب إجازة')
    flash('تم إرسال طلب الإجازة', 'success')
    return redirect(url_for('leaves'))

@app.route('/leaves/<int:leave_id>/<status>', methods=['POST'])
@admin_required
@csrf_required
def leave_status(leave_id, status):
    if status not in ('approved','rejected'):
        abort(400)
    db.update_leave_status(leave_id, status, session.get('username',''))
    db.log_audit('leave_status', 'leave', leave_id, f'{ "اعتماد" if status=="approved" else "رفض" } إجازة')
    flash(f'تم { "اعتماد" if status=="approved" else "رفض" } الإجازة', 'success')
    return redirect(url_for('leaves'))

# ==================== OVERTIME ====================
@app.route('/overtime')
@login_required
def overtime():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    emps = db.get_employees(status='active')
    records = db.get_overtime(month)
    return render_template('overtime.html', emps=emps, records=records, month=month)

@app.route('/overtime/save', methods=['POST'])
@role_required('user')
@csrf_required
def save_overtime():
    month = request.json.get('month')
    db.bulk_save_overtime(month, request.json.get('records', []))
    db.log_audit('overtime', 'overtime', None, f'تحديث إضافي شهر {month}')
    return jsonify({'ok': True})

# ==================== COMMISSIONS ====================
@app.route('/commissions')
@login_required
def commissions():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    emps = db.get_employees(status='active')
    records = db.get_commissions(month)
    return render_template('commissions.html', emps=emps, records=records, month=month)

@app.route('/commissions/save', methods=['POST'])
@role_required('user')
@csrf_required
def save_commissions():
    month = request.json.get('month')
    db.bulk_save_commissions(month, request.json.get('records', []))
    db.log_audit('commissions', 'commissions', None, f'تحديث عمولات شهر {month}')
    return jsonify({'ok': True})

# ==================== DEDUCTIONS ====================
@app.route('/deductions')
@login_required
def deductions():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    emps = db.get_employees(status='active')
    records = db.get_deductions(month)
    types = db.get_deduction_types()
    return render_template('deductions.html', emps=emps, records=records, types=types, month=month)

@app.route('/deductions/save', methods=['POST'])
@role_required('user')
@csrf_required
def save_deductions():
    month = request.json.get('month')
    db.bulk_save_deductions(month, request.json.get('records', []))
    db.log_audit('deductions', 'deductions', None, f'تحديث خصومات شهر {month}')
    return jsonify({'ok': True})

# ==================== BULK INCENTIVES (جماعية) ====================
# حوافز/خصومات تُطبق على الكل دفعة واحدة -> تُسجَّل كحوافز إضافية
@app.route('/bulk-incentives', methods=['GET','POST'])
@role_required('supervisor')
@csrf_required
def bulk_incentives():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    if request.method == 'POST':
        d = request.form.to_dict()
        d['month'] = (request.form.get('month') or datetime.now().strftime('%Y-%m'))
        d['kind'] = request.form.get('kind','bonus')
        d['per_employee_amount'] = _num(request.form.get('per_employee_amount'))
        spe = []
        for key, val in request.form.items():
            if key.startswith('emp_') and val == 'on':
                spe.append(key.replace('emp_',''))
        d['specific_employees'] = json.dumps(spe, ensure_ascii=False)
        db.add_bulk_incentive(d, session.get('username',''))
        db.apply_bulk_to_payroll(d['month'])
        db.log_audit('bulk_incentive', 'incentive', None, f'حافز/خصم جماعي شهر {d["month"]}')
        flash('تم تطبيق الحافز/الخصم الجماعي', 'success')
        return redirect(url_for('bulk_incentives', month=month))
    emps = db.get_employees(status='active')
    items = db.get_bulk_incentives(month)
    return render_template('bulk_incentives.html', emps=emps, items=items, month=month)

@app.route('/bulk-incentives/<int:bid>/delete', methods=['POST'])
@login_required
@csrf_required
def delete_bulk_incentive(bid):
    db.delete_bulk_incentive(bid)
    db.log_audit('delete', 'incentive', bid, 'حذف حافز جماعي')
    flash('تم الحذف', 'success')
    return redirect(url_for('bulk_incentives'))

# ==================== ADVANCES ====================
@app.route('/advances')
@login_required
def advances():
    emps = db.get_employees(status='active')
    records = db.get_advances()
    return render_template('advances.html', emps=emps, records=records)

@app.route('/advances/add', methods=['POST'])
@role_required('supervisor')
@csrf_required
def add_advance():
    d = request.form.to_dict()
    d['amount'] = _num(request.form.get('amount'))
    d['installment_count'] = int(_num(request.form.get('installment_count'), 1))
    db.add_advance(d)
    db.log_audit('add', 'advance', None, 'إضافة سلفة')
    flash('تم إضافة السلفة', 'success')
    return redirect(url_for('advances'))

@app.route('/advances/<int:adv_id>/delete', methods=['POST'])
@admin_required
@csrf_required
def delete_advance(adv_id):
    db.delete_advance(adv_id)
    db.log_audit('delete', 'advance', adv_id, 'حذف سلفة')
    flash('تم حذف السلفة', 'success')
    return redirect(url_for('advances'))

# ==================== PAYROLL ====================
@app.route('/payroll')
@login_required
def payroll():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    calculated = db.calculate_payroll(month)
    saved = db.get_saved_payroll(month)
    settings = db.get_settings()
    signs = db.get_signatures(month)
    signed_map = {s['employee_id']: s for s in signs}
    return render_template('payroll.html', data=calculated, saved=saved, month=month,
                           settings=settings, signed_map=signed_map)

@app.route('/payroll/save', methods=['POST'])
@role_required('supervisor')
@csrf_required
def save_payroll():
    month = request.json.get('month')
    db.save_payroll(month, request.json.get('records', []))
    db.log_audit('payroll_save', 'payroll', None, f'حفظ كشف مرتبات شهر {month}')
    return jsonify({'ok': True})

@app.route('/payroll/confirm', methods=['POST'])
@role_required('supervisor')
@csrf_required
def confirm_payroll():
    month = request.form['month']
    db.confirm_payroll(month, session.get('username',''))
    db.log_audit('payroll_confirm', 'payroll', None, f'اعتماد صرف مرتبات شهر {month}')
    flash('تم اعتماد صرف المرتبات', 'success')
    return redirect(url_for('payroll', month=month))

@app.route('/payroll/export', methods=['POST'])
@role_required('supervisor')
@csrf_required
def export_payroll():
    month = request.form['month']
    data = db.calculate_payroll(month)
    settings = db.get_settings()
    wb = Workbook()
    ws = wb.active
    ws.title = f"مرتبات {month}"
    ws.merge_cells('A1:Q1')
    ws['A1'] = f"{settings['company_name']} - كشف مرتبات شهر {month}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    headers = ['الكود','الاسم','القسم','الوظيفة','الأساسي','البدلات','إضافي','عمولة','إجمالي المستحق',
               'تأمينات(موظف)','ضرائب','سلف','غياب','تأخير','خصومات أخرى','إجمالي الخصومات','صافي الراتب']
    ws.append([])
    ws.append(headers)
    rows = [[r['emp_code'],r['name'],r['dept_name'],r['position'],r['base'],r['allowances'],
             r['overtime_amt'],r['commissions'],r['gross'],r['si_employee'],r['income_tax'],
             r['advance_installments'],r['absence_deduction'],r['late_deduction'],r['other_deductions'],
             r['total_deductions'],r['net']] for r in data]
    for row in rows:
        ws.append(row)
    total_row = len(rows) + 4
    ws.cell(row=total_row, column=2, value='الإجمالي')
    total_cols = [5,6,7,8,9,10,11,12,13,14,15,16,17]
    # صف الإجمالي: تجميع الأعمدة المالية فقط
    for col in total_cols:
        total = sum(row[col-1] for row in rows)
        ws.cell(row=total_row, column=col, value=round(total,2))
    _style_sheet(ws, header_row=3, ncols=len(headers),
                 total_rows=(total_row,),
                 money_cols=[5,6,7,8,9,10,11,12,13,14,15,16,17],
                 center_cols=[1,3,4])
    out = _export_xlsx(wb)
    return send_file(out, as_attachment=True, download_name=f'مرتبات-{month}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/payslip/<int:emp_id>/<month>')
@login_required
def payslip(emp_id, month):
    emp = db.get_employee(emp_id)
    if not emp:
        abort(404)
    calc = db.calculate_payroll(month)
    record = next((r for r in calc if r['employee_id']==emp_id), None)
    settings = db.get_settings()
    signed = db.has_signed(emp_id, month)
    return render_template('payslip.html', emp=emp, rec=record, month=month, settings=settings, signed=signed)

@app.route('/payslip/<int:emp_id>/<month>/sign', methods=['POST'])
@login_required
@csrf_required
def sign_payslip(emp_id, month):
    emp = db.get_employee(emp_id)
    if not emp:
        abort(404)
    calc = db.calculate_payroll(month)
    record = next((r for r in calc if r['employee_id']==emp_id), None)
    if not record:
        flash('لا يوجد كشف مرتب لهذا الموظف', 'error')
        return redirect(url_for('payslip', emp_id=emp_id, month=month))
    sig = request.form.get('signature_text') or request.form.get('signature_data') or ''
    # صورة توقيع اختيارية
    sig_img = ''
    if 'signature_img' in request.files and request.files['signature_img'].filename:
        f = request.files['signature_img']
        safe = f.read()
        if len(safe) > 300*1024:
            flash('حجم صورة التوقيع كبير', 'error')
            return redirect(url_for('payslip', emp_id=emp_id, month=month))
        import base64
        sig_img = base64.b64encode(safe).decode('ascii')
    sign_data = sig if sig else (f"img:{sig_img}" if sig_img else '')
    db.sign_payroll(emp_id, month, f"SLIP-{emp_id}-{month}", record['gross'], record['net'],
                    sign_data, request.remote_addr or '', request.form.get('notes',''))
    db.log_audit('sign', 'payslip', emp_id, f'توقيع استلام راتب {month}')
    flash('تم تسجيل توقيع الاستلام بنجاح', 'success')
    return redirect(url_for('payslip', emp_id=emp_id, month=month))

@app.route('/payslip/<int:emp_id>/<month>/pdf')
@login_required
def payslip_pdf(emp_id, month):
    emp = db.get_employee(emp_id)
    if not emp:
        abort(404)
    calc = db.calculate_payroll(month)
    record = next((r for r in calc if r['employee_id']==emp_id), None)
    settings = db.get_settings()
    return _build_payslip_pdf(emp, record, month, settings)

# ==================== CONTRACTS ====================
@app.route('/contracts')
@login_required
def contracts():
    emps = db.get_employees(status='all')
    records = db.get_contracts()
    return render_template('contracts.html', emps=emps, records=records)

@app.route('/contracts/add', methods=['POST'])
@role_required('supervisor')
@csrf_required
def add_contract():
    d = request.form.to_dict()
    d['salary'] = _num(request.form.get('salary'))
    new_id = db.add_contract(d)
    db.log_audit('add', 'contract', new_id, 'إضافة عقد')
    flash('تم إضافة العقد', 'success')
    return redirect(url_for('contracts'))

@app.route('/contracts/<int:cid>/edit', methods=['POST'])
@login_required
@csrf_required
def edit_contract(cid):
    d = request.form.to_dict()
    d['salary'] = _num(request.form.get('salary'))
    db.update_contract(cid, d)
    db.log_audit('edit', 'contract', cid, 'تعديل عقد')
    flash('تم تحديث العقد', 'success')
    return redirect(url_for('contracts'))

@app.route('/contracts/<int:cid>/delete', methods=['POST'])
@admin_required
@csrf_required
def delete_contract(cid):
    db.delete_contract(cid)
    db.log_audit('delete', 'contract', cid, 'حذف عقد')
    flash('تم حذف العقد', 'success')
    return redirect(url_for('contracts'))

# ==================== PAYROLL COMPARISON (مقارنة شهرين) ====================
@app.route('/reports/comparison')
@login_required
def comparison():
    m1 = request.args.get('month1', datetime.now().strftime('%Y-%m'))
    m2 = request.args.get('month2', (datetime.now().replace(day=1) - timedelta(days=1)).strftime('%Y-%m'))
    data1 = {r['employee_id']: r for r in db.calculate_payroll(m1)}
    data2 = {r['employee_id']: r for r in db.calculate_payroll(m2)}
    ids = sorted(set(data1.keys()) | set(data2.keys()))
    rows = []
    for eid in ids:
        a = data1.get(eid); b = data2.get(eid)
        if not a: a = {k: 0 for k in (b or {})}; a['name']=b['name']; a['emp_code']=b['emp_code']; a['employee_id']=eid
        if not b: b = {k: 0 for k in (a or {})}; b['name']=a['name']; b['emp_code']=a['emp_code']; b['employee_id']=eid
        rows.append({'id': eid, 'name': a.get('name',''), 'emp_code': a.get('emp_code',''),
                     'net1': a.get('net',0), 'net2': b.get('net',0),
                     'diff': round((b.get('net',0) or 0) - (a.get('net',0) or 0), 2),
                     'gross1': a.get('gross',0), 'gross2': b.get('gross',0),
                     'ded1': a.get('total_deductions',0), 'ded2': b.get('total_deductions',0)})
    return render_template('comparison.html', rows=rows, m1=m1, m2=m2)

# ==================== REPORTS ====================
@app.route('/reports')
@login_required
def reports():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    data = db.calculate_payroll(month)
    signatures = db.get_signatures(month)
    return render_template('reports.html', data=data, month=month, signatures=signatures)

@app.route('/reports/department')
@login_required
def dept_report():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    data = db.calculate_payroll(month)
    from collections import defaultdict
    grouped = defaultdict(lambda: {'count':0,'base':0,'earnings':0,'deductions':0,'net':0})
    for r in data:
        dept = r['dept_name'] or 'بدون قسم'
        g = grouped[dept]
        g['count'] += 1
        g['base'] += r['base']
        g['earnings'] += r['gross']
        g['deductions'] += r['total_deductions']
        g['net'] += r['net']
    return render_template('dept_report.html', grouped=dict(grouped), month=month)

@app.route('/reports/department/export')
@login_required
def dept_report_export():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    from collections import defaultdict
    grouped = defaultdict(lambda: {'count':0,'base':0,'earnings':0,'deductions':0,'net':0})
    for r in db.calculate_payroll(month):
        dept = r['dept_name'] or 'بدون قسم'
        g = grouped[dept]
        g['count'] += 1; g['base'] += r['base']; g['earnings'] += r['gross']
        g['deductions'] += r['total_deductions']; g['net'] += r['net']
    wb = Workbook(); ws = wb.active; ws.title = "تقرير الأقسام"
    headers = ['القسم','عدد الموظفين','الأساسي','إجمالي المستحق','الإجمالي للخصومات','صافي']
    ws.append(headers)
    for dept, g in grouped.items():
        ws.append([dept, g['count'], g['base'], g['earnings'], g['deductions'], g['net']])
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value='الإجمالي')
    for col, key in [(3,'base'),(4,'earnings'),(5,'deductions'),(6,'net')]:
        ws.cell(row=total_row, column=col, value=round(sum(g[k] for g in grouped.values()), 2))
    ws.cell(row=total_row, column=2, value=sum(g['count'] for g in grouped.values()))
    _style_sheet(ws, 1, len(headers), total_rows=(total_row,),
                 money_cols=[3,4,5,6], center_cols=[2])
    out = _export_xlsx(wb)
    return send_file(out, as_attachment=True, download_name=f'تقرير-الأقسام-{month}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/reports/annual')
@login_required
def annual_report():
    year = request.args.get('year', str(datetime.now().year))
    result = {}
    grand = {'net':0,'earnings':0,'deductions':0}
    for m in range(1,13):
        month = f"{year}-{m:02d}"
        data = db.calculate_payroll(month)
        result[month] = {'count': len(data), 'net': sum(r['net'] for r in data),
                         'earnings': sum(r['gross'] for r in data),
                         'deductions': sum(r['total_deductions'] for r in data)}
        grand['net'] += result[month]['net']; grand['earnings'] += result[month]['earnings']
        grand['deductions'] += result[month]['deductions']
    return render_template('annual_report.html', data=result, year=year, grand=grand)

# ==================== SIGNATURES LOG (سجل التوقيعات) ====================
@app.route('/reports/signatures')
@login_required
def signatures_log():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    signs = db.get_signatures(month)
    return render_template('signatures.html', signs=signs, month=month)

# ==================== AUDIT LOG (سجل التدقيق) ====================
@app.route('/audit')
@admin_required
def audit():
    day = request.args.get('day', '')
    if day:
        logs = db.get_audit_log(1000, day=day)
    else:
        logs = db.get_audit_log(300)
    return render_template('audit.html', logs=logs, day=day)

@app.route('/audit/export')
@admin_required
def audit_export():
    day = request.args.get('day', '')
    logs = db.get_audit_log(10000, day=day)
    wb = Workbook(); ws = wb.active; ws.title = "سجل التدقيق"
    headers = ['التاريخ','المستخدم','الإجراء','الكيان','التفاصيل','IP']
    ws.append(headers)
    for l in logs:
        ws.append([l['created_at'], l['username'], l['action'], l['entity'], l['details'], l['ip_address']])
    _style_sheet(ws, 1, len(headers), center_cols=[1,2,3,4,6])
    out = _export_xlsx(wb)
    fname = f'سجل-التدقيق-{day}.xlsx' if day else 'سجل-التدقيق.xlsx'
    return send_file(out, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ==================== SETTINGS ====================
@app.route('/settings', methods=['GET','POST'])
@login_required
@csrf_required
def settings():
    if request.method == 'POST':
        if session.get('role') != 'admin':
            flash('الإعدادات تعديلها للمسؤول فقط', 'error')
            return redirect(url_for('settings'))
        db.save_settings(request.form.to_dict())
        db.log_audit('settings', 'settings', None, 'تحديث الإعدادات')
        flash('تم حفظ الإعدادات', 'success')
        return redirect(url_for('settings'))
    s = db.get_settings()
    brackets = db.get_tax_brackets()
    backups = db.list_backups()
    return render_template('settings.html', s=s, brackets=brackets, backups=backups)

# ==================== BACKUP (نسخ احتياطي) ====================
@app.route('/settings/backup', methods=['POST'])
@admin_required
@csrf_required
def backup_create():
    try:
        path, when = db.create_backup()
        db.log_audit('backup', 'system', None, f'نسخة احتياطية يدوية {when}')
        flash(f'تم إنشاء النسخة الاحتياطية', 'success')
    except Exception as e:
        flash(f'فشل إنشاء النسخة الاحتياطية: {e}', 'error')
    return redirect(url_for('settings'))

@app.route('/settings/backups/<path:filename>/restore', methods=['POST'])
@admin_required
@csrf_required
def backup_restore(filename):
    try:
        db.restore_backup(filename)
        db.log_audit('restore', 'system', None, f'استعادة نسخة احتياطية {filename}')
        flash('تمت الاستعادة بنجاح. تم تسجيل الخروج من الجلسة.', 'success')
        session.clear()
        return redirect(url_for('login'))
    except Exception as e:
        flash(f'فشل الاستعادة: {e}', 'error')
        return redirect(url_for('settings'))

@app.route('/settings/backups/<path:filename>/delete', methods=['POST'])
@admin_required
@csrf_required
def backup_delete(filename):
    if db.delete_backup(filename):
        db.log_audit('backup_delete', 'system', None, f'حذف نسخة احتياطية {filename}')
        flash('تم حذف النسخة الاحتياطية', 'success')
    else:
        flash('تعذر حذف النسخة الاحتياطية', 'error')
    return redirect(url_for('settings'))

@app.route('/settings/backups/<path:filename>/download')
@admin_required
def backup_download(filename):
    if not (filename.startswith('backup_') and filename.endswith('.db')):
        abort(404)
    p = os.path.join(os.path.join(DATA_DIR, 'backups'), filename)
    if not os.path.isfile(p):
        abort(404)
    with open(p, 'rb') as f:
        data = f.read()
    return send_file(io.BytesIO(data), as_attachment=True, download_name=filename,
                     mimetype='application/octet-stream')

# ==================== USERS ====================
@app.route('/users')
@admin_required
def users():
    all = db.get_all_users()
    return render_template('users.html', users=all)

@app.route('/users/add', methods=['POST'])
@admin_required
@csrf_required
def add_user():
    u = request.form.to_dict()
    if len(u.get('password','')) < 8:
        flash('كلمة المرور يجب أن تكون 8 أحرف على الأقل', 'error')
        return redirect(url_for('users'))
    # تحقق من صلاحية صالحة فقط
    valid_roles = ('admin','supervisor','user','viewer')
    role = u.get('role','user')
    if role not in valid_roles:
        role = 'user'
    u['role'] = role
    try:
        db.add_user(u)
        db.log_audit('add', 'user', None, f'إضافة مستخدم: {u.get("username")} ({role})')
        flash('تم إضافة المستخدم', 'success')
    except Exception:
        flash('اسم المستخدم موجود بالفعل', 'error')
    return redirect(url_for('users'))

@app.route('/users/<int:uid>/toggle', methods=['POST'])
@admin_required
@csrf_required
def toggle_user(uid):
    if uid == session.get('user_id'):
        flash('لا يمكنك تعطيل حسابك', 'error')
    else:
        db.toggle_user(uid)
        db.log_audit('toggle', 'user', uid, 'تفعيل/تعطيل مستخدم')
        flash('تم تحديث حالة المستخدم', 'success')
    return redirect(url_for('users'))

@app.route('/users/<int:uid>/delete', methods=['POST'])
@admin_required
@csrf_required
def delete_user(uid):
    if uid == session.get('user_id'):
        flash('لا يمكنك حذف حسابك', 'error')
    else:
        db.delete_user(uid)
        db.log_audit('delete', 'user', uid, 'حذف مستخدم')
        flash('تم حذف المستخدم', 'success')
    return redirect(url_for('users'))

# ==================== GUIDE (دليل الاستخدام) ====================
@app.route('/guide')
@login_required
def guide():
    return render_template('guide.html')

# ==================== PDF generation ====================
def _build_payslip_pdf(emp, record, month, settings):
    if record is None:
        from flask import abort
        abort(404)
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib.utils import ImageReader
        from reportlab.lib.colors import HexColor
        import arabic_reshaper
        from bidi.algorithm import get_display
        from PIL import Image as PILImage
    except Exception as e:
        flash(f'مكتبة PDF غير متوفرة: {e}', 'error')
        return redirect(url_for('payslip', emp_id=emp['id'], month=month))

    # تحميل خط عربي
    font_name = 'Arial'
    font_bold = 'ArialBold'
    font_path = r'C:\Windows\Fonts\arial.ttf'
    font_bold_path = r'C:\Windows\Fonts\arialbd.ttf'
    if not os.path.exists(font_path):
        font_path = os.path.join(BASE_DIR, 'static', 'fonts', 'arial.ttf')
    try:
        pdfmetrics.registerFont(TTFont(font_name, font_path))
        if os.path.exists(font_bold_path):
            try:
                pdfmetrics.registerFont(TTFont(font_bold, font_bold_path))
            except Exception:
                font_bold = font_name
        else:
            font_bold = font_name
    except Exception:
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'

    def ar(text):
        try:
            return get_display(arabic_reshaper.reshape(str(text)))
        except Exception:
            return str(text)

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    margin = 40
    cy = H - 45

    def heading(txt, size, y, color=None):
        fnt = font_bold if font_bold in pdfmetrics.getRegisteredFontNames() else font_name
        c.setFont(fnt, size)
        c.setFillColor(color or HexColor('#1e3a5f'))
        c.drawCentredString(W/2, y, ar(txt))
        return y - size - 10

    def row_label_value(label, value, x1, x2, y, size=10):
        c.setFont(font_bold if font_bold in pdfmetrics.getRegisteredFontNames() else font_name, size)
        c.setFillColor(HexColor('#4b5563'))
        c.drawString(x1, y, ar(label))
        c.setFillColor(HexColor('#111827'))
        c.drawRightString(x2, y, ar(value))

    # header company
    if settings.get('company_name'):
        cy = heading(settings['company_name'], 15, cy)
    cy = heading('إيصال استلام المرتب', 13, cy, HexColor('#2563eb'))
    cy -= 4
    c.setStrokeColor(HexColor('#2563eb')); c.setLineWidth(1.2)
    c.line(margin, cy, W-margin, cy)
    cy -= 18

    # slip info
    c.setFont(font_bold if font_bold in pdfmetrics.getRegisteredFontNames() else font_name, 9)
    c.setFillColor(HexColor('#6b7280'))
    c.drawString(margin, cy, ar(f"رقم القسيمة: SLIP-{emp['emp_code']}-{month}    الشهر: {month}"))
    cy -= 18

    # employee data
    c.setFont(font_bold, 10); c.setFillColor(HexColor('#1e3a5f'))
    c.drawString(margin, cy, ar('بيانات الموظف'))
    cy -= 16
    grid_items = [('الاسم', emp['name']), ('الكود', emp['emp_code']), ('القسم', emp.get('dept_name','')),
                  ('الوظيفة', emp['position']), ('نوع الأجر', emp['pay_type'])]
    kw = 2; kws = (W - 2*margin) / (kw*2)
    for i in range(0, len(grid_items), 2):
        row = grid_items[i:i+2]
        x = margin
        for label, val in row:
            row_label_value(label+': ', val, x, x+kws, cy, 9)
            x += kws
        cy -= 16

    cy -= 6

    # earnings
    c.setFont(font_bold, 10); c.setFillColor(HexColor('#1e3a5f'))
    c.drawString(margin, cy, ar('المستحقات'))
    cy -= 14
    c.setFont(font_name, 9); c.setFillColor(HexColor('#111827'))
    def money_line(label, val, y):
        c.drawString(margin+24, y, ar(label))
        c.drawRightString(W-margin, y, ar(f"{val:,.2f} {settings.get('currency','')}"))
    money_line('الأساسي', record['base'], cy); cy -= 14
    money_line('إجمالي البدلات', record['allowances'], cy); cy -= 14
    money_line('الإضافي', record['overtime_amt'], cy); cy -= 14
    money_line('العمولات', record['commissions'], cy); cy -= 14
    c.setFont(font_bold,10); c.setFillColor(HexColor('#16a34a'))
    money_line('إجمالي المستحق', record['gross'], cy); cy -= 18

    # deductions
    c.setFont(font_bold, 10); c.setFillColor(HexColor('#1e3a5f'))
    c.drawString(margin, cy, ar('الخصومات'))
    cy -= 14
    c.setFont(font_name, 9); c.setFillColor(HexColor('#111827'))
    money_line('تأمينات (موظف)', record['si_employee'], cy); cy -= 14
    money_line('ضريبة كسب العمل', record['income_tax'], cy); cy -= 14
    money_line('أقساط سلف', record['advance_installments'], cy); cy -= 14
    money_line('خصم غياب', record['absence_deduction'], cy); cy -= 14
    money_line('خصم تأخير', record['late_deduction'], cy); cy -= 14
    money_line('خصومات أخرى', record['other_deductions'], cy); cy -= 14
    c.setFont(font_bold,10); c.setFillColor(HexColor('#dc2626'))
    money_line('إجمالي الخصومات', record['total_deductions'], cy); cy -= 20

    # net salary box
    c.setFillColor(HexColor('#2563eb'))
    c.roundRect(margin, cy-52, W-2*margin, 44, 10, fill=1)
    c.setFillColor(HexColor('#ffffff'))
    c.setFont(font_bold, 11)
    c.drawString(margin+20, cy-22, ar('صافي الراتب المستحق: '))
    c.drawRightString(W-margin-20, cy-22, ar(f"{record['net']:,.2f} {settings.get('currency','')}"))
    cy -= 66

    # signature/stamp area
    c.setStrokeColor(HexColor('#cbd5e1'))
    c.setFont(font_name, 9); c.setFillColor(HexColor('#6b7280'))
    # signature taken?
    try:
        signed = db.has_signed(emp['id'], month)
    except Exception:
        signed = False
    if signed:
        c.drawString(margin, cy, ar('✓ تم التوقيع على الاستلام إلكترونياً'))
    c.drawString(margin, cy-30, ar('توقيع الموظف: ____________'))
    c.drawRightString(W-margin, cy-30, ar('ختم الشركة وتوقيع مسؤول الصرف: ____________'))
    cy -= 55
    c.setFont(font_name, 8); c.setFillColor(HexColor('#9ca3af'))
    c.drawCentredString(W/2, 30, ar('نظام المرتبات الشامل — جميع الحقوق محفوظة © 2026 محاسب / أحمد عبد الله'))

    c.showPage()
    c.save()
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'قسيمة-مكان-{emp["emp_code"]}-{month}.pdf',
                     mimetype='application/pdf')

# ==================== Helpers ====================
def _num(v, default=0):
    try:
        return float(str(v or '0').replace(',', '').strip() or '0')
    except Exception:
        return default

# ==================== RUN ====================
import webbrowser, threading

# ==================== ورقة Excel فخمة (تصدير موحّد) ====================
_HEADER_FILL = PatternFill('solid', fgColor='1F3864')     # أزرق داكن
_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
_ALT_FILL = PatternFill('solid', fgColor='DCE6F1')         # تظليل بديل فاتح
_TOT_FILL = PatternFill('solid', fgColor='FFE699')         # خبرة الإجمالي أصفر
_THIN = Side(style='thin', color='9CA3AF')
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal='center', vertical='center')
_RIGHT = Alignment(horizontal='right', vertical='center')

def _style_sheet(ws, header_row, ncols, total_rows=(), money_cols=(), center_cols=()):
    """تطبيق شكل فخم على كل ورقة تقرير: ترويسة مسطحة + حدود + تظليل متناوب + عرض أعمدة + تجميد + فلتر."""
    from openpyxl.utils import get_column_letter as _gcl
    # ترويسة
    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BOX
    # بيانات (حدود + تظليل متناوب + محاذاة)
    for row in range(header_row + 1, ws.max_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = _BOX
            if row in total_rows:
                cell.fill = _TOT_FILL
                cell.font = Font(bold=True, size=11)
            elif (row - header_row) % 2 == 0:
                cell.fill = _ALT_FILL
            if c in money_cols:
                cell.number_format = '#,##0.00'
                cell.alignment = _RIGHT
            elif c in center_cols:
                cell.alignment = _CENTER
    # عرض أعمدة مناسب
    for c in range(1, ncols + 1):
        letter = _gcl(c)
        col_vals = [ws.cell(row=r, column=c).value for r in range(header_row, ws.max_row + 1)]
        width = max((len(str(v)) if v is not None else 0) for v in col_vals) + 4 if col_vals else 12
        ws.column_dimensions[letter].width = min(max(width, 10), 32)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if ws.max_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:{_gcl(ncols)}{ws.max_row}"
    ws.row_dimensions[header_row].height = 22

def _export_xlsx(wb):
    """تحويل المصنف إلى استجابة تنزيل."""
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def _open_browser(port):
    time.sleep(1.5)
    try:
        webbrowser.open(f'http://localhost:{port}/login')
    except Exception:
        pass

def run_server():
    # التشغيل الآمن: debug معطّل في الإنتاج
    debug = os.environ.get('PAYROLL_DEBUG', '0') == '1'
    port = int(os.environ.get('PORT', 5000))
    if os.environ.get('PAYROLL_NO_BROWSER', '0') != '1':
        threading.Thread(target=_open_browser, args=(port,), daemon=True).start()
    app.run(debug=debug, host='0.0.0.0', port=port, threaded=True)

if __name__ == '__main__':
    run_server()

